"""
EGTS Bidirectional Excel Parser
================================

Двунаправленный разбор EGTS ↔ Excel:
  DECODE: binary/hex/csv → .xlsx  (заполняет редактируемые ячейки + HEX-поля)
  ENCODE: .xlsx → binary          (читает атрибуты, пересчитывает CRC/длины, пишет .bin + обновляет HEX в .xlsx)

Использование:
  python egts_excel_parser.py decode <input.csv>  [output.xlsx]
  python egts_excel_parser.py decode --hex <HEX>  [output.xlsx]
  python egts_excel_parser.py encode <input.xlsx> [output.bin]
  python egts_excel_parser.py roundtrip <input.xlsx>   # re-encode + обновить HEX в xlsx

Цветовая схема:
  Жёлтый фон   — редактируемые атрибуты (меняйте эти ячейки)
  Синий фон     — вычисляемые HEX-поля (обновляются при encode)
  Зелёный фон   — валидные CRC
  Красный фон   — ошибки / невалидные CRC

Листы:
  LEGEND          — инструкция по использованию
  PACKETS         — один пакет = одна строка, ключевые поля + HEX пакета
  HEADER          — поля заголовка (редактируемые)
  SDR             — Service Data Records
  SRT_POS_DATA    — EGTS_SR_POS_DATA (SRT=16)
  SRT_EXT_POS     — EGTS_SR_EXT_POS_DATA (SRT=17)
  SRT_STATE       — EGTS_SR_STATE_DATA (SRT=20/21)
  SRT_LIQUID      — EGTS_SR_LIQUID_LEVEL_SENSOR (SRT=27)
  SRT_AD_SENSORS  — EGTS_SR_AD_SENSORS_DATA (SRT=18)
  SRT_ABS_CNTR    — EGTS_SR_ABS_CNTR_DATA (SRT=25)
  SRT_TERM_IDENT  — EGTS_SR_TERM_IDENTITY (SRT=1)
  SRT_200         — RTLS Extended Position
  SRT_201         — RTLS Sensor Data
  SRT_202         — RTLS Tag Identity
  SRT_203         — RTLS Event Data
  SRT_205         — LBS (base stations) data for road graph positioning (new, discussion 18)
  SRT_RAW         — нераспознанные подзаписи
"""

import json
import struct
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите: pip install openpyxl")
    sys.exit(1)

# SERVICE в пути
_svc = str(Path(__file__).parent.parent / "SERVICE")
if _svc not in sys.path:
    sys.path.insert(0, _svc)

from egts.codec import parse_stream, build_packet, EGTSPacket, ServiceDataRecord, RecordData, PtResponse
from egts.models import (
    SrPosData, SrExtPosData, SrStateData, SrLiquidLevelSensor,
    SrAdSensorsData, SrAbsCntrData, SrTermIdentity,
    SrRecordResponse, SrResultCode, SrAuthInfo,
    SrCountersData, SrPassengersCounters,
    SrCustom200, SrCustom201, SrCustom202, SrCustom203, SrCustom204, SrCustom205, SrRaw,
)
from egts.const import SRT_NAMES, PT_NAMES, SVC_NAMES, RESULT_CODES


# ---------------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------------

YELLOW  = "FFF2CC"   # редактируемые атрибуты
BLUE    = "DEEAF1"   # вычисляемые HEX
GREEN   = "E2EFDA"   # OK / valid
RED     = "FCE4D6"   # ERROR
GRAY    = "F2F2F2"   # чётные строки
HEADER  = "1F4E79"   # шапка таблицы
HEADER2 = "2E75B6"   # подзаголовок секции
SRT200C = "FFF2CC"
SRT201C = "E2EFDA"
SRT202C = "DEEAF1"
SRT203C = "FCE4D6"
SRT204C = "EAD1DC"  # IMU / Inertial (discussion 09, 13-16)
SRT205C = "D9EAD3"  # LBS (discussion 18)


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _style_header(ws, row: int, ncols: int, color: str = HEADER) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(color)
        c.font = _font(bold=True, color="FFFFFF")
        c.alignment = _center()
        c.border = _border()


def _style_attr(ws, row: int, ncols: int, bg: str | None = None) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        if bg:
            c.fill = _fill(bg)
        c.alignment = _left()
        c.border = _border()
        c.font = _font(size=9)


def _style_hex_cell(cell) -> None:
    cell.fill = _fill(BLUE)
    cell.font = _font(size=8, color="1F4E79")
    cell.alignment = _left()
    cell.border = _border()


def _auto_width(ws) -> None:
    for col in ws.columns:
        ml = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 2, 12), 60)


def _title(ws, text: str, ncols: int, color: str = HEADER2) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws["A1"]
    c.value = text
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.fill = _fill(color)
    c.alignment = _center()
    ws.row_dimensions[1].height = 22


def _headers(ws, cols: list[str], row: int = 2, color: str = HEADER) -> None:
    for i, h in enumerate(cols, 1):
        ws.cell(row=row, column=i).value = h
    _style_header(ws, row, len(cols), color)


# Метки типа колонки в строке 3
_ATTR_LABEL = "ATTR"    # редактируемая
_HEX_LABEL  = "HEX"    # вычисляемая


def _col_labels(ws, types: list[str], row: int = 3) -> None:
    for col, t in enumerate(types, 1):
        c = ws.cell(row=row, column=col)
        c.value = t
        if t == _ATTR_LABEL:
            c.fill = _fill(YELLOW)
            c.font = _font(bold=True, color="7F6000", size=8)
        elif t == _HEX_LABEL:
            c.fill = _fill(BLUE)
            c.font = _font(bold=True, color="1F4E79", size=8)
        else:
            c.fill = _fill(GRAY)
            c.font = _font(size=8)
        c.alignment = _center()
        c.border = _border()


# ---------------------------------------------------------------------------
# Контекст строки
# ---------------------------------------------------------------------------

class _Row:
    """Хранит связь строки листа с пакетом/SDR/SR."""
    def __init__(self, pkt_idx: int, sdr_idx: int, sr_idx: int, pkt: EGTSPacket, sdr=None, rd=None):
        self.pkt_idx = pkt_idx
        self.sdr_idx = sdr_idx
        self.sr_idx  = sr_idx
        self.pkt     = pkt
        self.sdr     = sdr
        self.rd      = rd


# ---------------------------------------------------------------------------
# DECODE: packets → workbook
# ---------------------------------------------------------------------------

def decode_to_workbook(packets: list[EGTSPacket]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_legend(wb)
    _sheet_packets(wb, packets)
    _sheet_header(wb, packets)
    _sheet_sdr(wb, packets)
    _sheet_pos_data(wb, packets)
    _sheet_ext_pos(wb, packets)
    _sheet_state(wb, packets)
    _sheet_liquid(wb, packets)
    _sheet_ad_sensors(wb, packets)
    _sheet_abs_cntr(wb, packets)
    _sheet_term_ident(wb, packets)
    _sheet_srt200(wb, packets)
    _sheet_srt201(wb, packets)
    _sheet_srt202(wb, packets)
    _sheet_srt203(wb, packets)
    _sheet_srt204(wb, packets)  # IMU / Inertial + fusion (discussion 09, 13-16)
    _sheet_srt205(wb, packets)  # LBS (discussion 18)
    _sheet_raw(wb, packets)
    return wb


def _collect(packets: list[EGTSPacket], srt_filter) -> list[_Row]:
    rows = []
    for pi, pkt in enumerate(packets):
        body = pkt.body if isinstance(pkt.body, list) else []
        for si, sdr in enumerate(body):
            for ri, rd in enumerate(sdr.record_data):
                if srt_filter is None or rd.srt == srt_filter or (isinstance(srt_filter, set) and rd.srt in srt_filter):
                    rows.append(_Row(pi, si, ri, pkt, sdr, rd))
    return rows


def _sheet_legend(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("LEGEND")
    ws.column_dimensions["A"].width = 80
    lines = [
        ("EGTS Bidirectional Excel Parser", True, HEADER),
        ("", False, None),
        ("DECODE: бинарный hex → атрибуты в Excel", False, None),
        ("ENCODE: атрибуты из Excel → бинарный пакет (пересчёт CRC и длин)", False, None),
        ("", False, None),
        ("Цветовая схема:", True, None),
        ("  Жёлтый фон — ATTR — редактируемые атрибуты пакета", False, YELLOW),
        ("  Синий фон  — HEX  — вычисляемые поля (обновляются при encode)", False, BLUE),
        ("  Зелёный    — OK / CRC валиден", False, GREEN),
        ("  Красный    — ERR / CRC не совпадает", False, RED),
        ("", False, None),
        ("Команды CLI:", True, None),
        ("  python egts_excel_parser.py decode packets.csv output.xlsx", False, None),
        ("  python egts_excel_parser.py decode --hex <HEX> output.xlsx", False, None),
        ("  python egts_excel_parser.py encode output.xlsx output.bin", False, None),
        ("  python egts_excel_parser.py roundtrip output.xlsx", False, None),
        ("", False, None),
        ("При encode:", True, None),
        ("  1. Читаются ATTR-ячейки (жёлтые) из всех листов", False, None),
        ("  2. Пересчитываются все длины (RL, FDL) и CRC (HCS, SFRCS)", False, None),
        ("  3. HEX-ячейки (синие) обновляются в xlsx", False, None),
        ("  4. Полный бинарный пакет записывается в .bin файл", False, None),
    ]
    for i, (text, bold, bg) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        if bg:
            c.fill = _fill(bg)
        ws.row_dimensions[i].height = 18


def _sheet_packets(wb, packets: list[EGTSPacket]) -> None:
    ws = wb.create_sheet("PACKETS")
    cols  = ["PKT#", "PT", "PT_name", "PID", "FDL", "SDR_count",
             "HCS_valid", "CRC16_valid", "PARSE_ERRORS",
             "HEX_PACKET"]                            # последняя — вычисляемая
    ctypes = ["", "ATTR","","ATTR","","","","","", "HEX"]
    _title(ws, "EGTS Packets", len(cols))
    _headers(ws, cols, row=2)
    _col_labels(ws, ctypes, row=3)

    for i, pkt in enumerate(packets):
        row = i + 4
        hdr = pkt.header
        body = pkt.body if isinstance(pkt.body, list) else []
        ok   = not pkt.parse_errors and pkt.hcs_valid and pkt.crc16_valid
        bg   = GREEN if ok else RED
        vals = [
            i + 1,
            hdr.packet_type,
            PT_NAMES.get(hdr.packet_type, f"PT_{hdr.packet_type}"),
            hdr.packet_id,
            hdr.frame_data_length,
            len(body),
            "OK" if pkt.hcs_valid else "FAIL",
            "OK" if pkt.crc16_valid else "FAIL",
            "; ".join(pkt.parse_errors) or "—",
            pkt.hex_raw,
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill = _fill(bg if col < len(cols) else BLUE)
            c.alignment = _left()
            c.border = _border()
            c.font = _font(size=9)
        _style_hex_cell(ws.cell(row=row, column=len(cols)))
    _auto_width(ws)


def _sheet_header(wb, packets: list[EGTSPacket]) -> None:
    ws  = wb.create_sheet("HEADER")
    # Все поля ATTR + HEX заголовка и CRC
    cols = ["PKT#", "PRV","SKID","PRF","RTE","ENA","CMP","PR",
            "HL","HE","FDL","PID","PT","PT_name",
            "PRA","RCA","TTL",
            "HCS","HCS_valid", "CRC16","CRC16_valid",
            "HEX_HEADER"]
    ct   = ["","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR",
            "HEX","ATTR","HEX","ATTR","ATTR","",
            "ATTR","ATTR","ATTR",
            "HEX","","HEX","",
            "HEX"]
    _title(ws, "EGTS Header Fields", len(cols))
    _headers(ws, cols, row=2)
    _col_labels(ws, ct, row=3)

    for i, pkt in enumerate(packets):
        row = i + 4
        h   = pkt.header
        hcs_ok = pkt.hcs_valid
        vals = [
            i+1,
            h.protocol_version, h.security_key_id,
            h.prefix, h.route, h.encryption_alg, h.compression, h.priority,
            h.header_length, h.header_encoding, h.frame_data_length,
            h.packet_id, h.packet_type,
            PT_NAMES.get(h.packet_type, f"PT_{h.packet_type}"),
            h.peer_address if h.routed else "",
            h.recipient_address if h.routed else "",
            h.time_to_live if h.routed else "",
            f"{h.header_crc:#04x}",
            "OK" if hcs_ok else "FAIL",
            f"{pkt.crc16_value:#06x}",
            "OK" if pkt.crc16_valid else "FAIL",
            pkt.hex_raw[:44] + "...",   # первые 22 байта (заголовок)
        ]
        row_bg = GRAY if i % 2 == 0 else None
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            tp = ct[col - 1]
            if tp == "ATTR":
                c.fill = _fill(YELLOW)
            elif tp == "HEX":
                _style_hex_cell(c); continue
            elif row_bg:
                c.fill = _fill(row_bg)
            c.alignment = _left()
            c.border = _border()
            c.font = _font(size=9)
    _auto_width(ws)


def _sheet_sdr(wb, packets: list[EGTSPacket]) -> None:
    ws  = wb.create_sheet("SDR")
    cols = ["PKT#","SDR#","RL","RN","SSOD","RSOD","GRP","RPP",
            "TMFE","EVFE","OBFE","OID","EVID","TM",
            "SST","SST_name","RST","RST_name",
            "SR_count","HEX_SDR"]
    ct   = ["","","HEX","ATTR","ATTR","ATTR","ATTR","ATTR",
            "ATTR","ATTR","ATTR","ATTR","ATTR","ATTR",
            "ATTR","","ATTR","",
            "","HEX"]
    _title(ws, "Service Data Records", len(cols))
    _headers(ws, cols, row=2)
    _col_labels(ws, ct, row=3)

    dr = 4
    for pi, pkt in enumerate(packets):
        body = pkt.body if isinstance(pkt.body, list) else []
        for si, sdr in enumerate(body):
            vals = [
                pi+1, si+1,
                sdr.record_length, sdr.record_number,
                sdr.ssod, sdr.rsod, sdr.grp, sdr.rpp,
                sdr.tmfe, sdr.evfe, sdr.obfe,
                sdr.object_id or "", sdr.event_id or "",
                sdr.record_time.isoformat() if sdr.record_time else "",
                sdr.source_service_type,
                SVC_NAMES.get(sdr.source_service_type, f"SVC_{sdr.source_service_type}"),
                sdr.recipient_service_type,
                SVC_NAMES.get(sdr.recipient_service_type, f"SVC_{sdr.recipient_service_type}"),
                len(sdr.record_data),
                sdr.to_bytes().hex().upper(),
            ]
            bg = GRAY if dr % 2 == 0 else None
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=dr, column=col, value=v)
                tp = ct[col-1]
                if tp == "ATTR": c.fill = _fill(YELLOW)
                elif tp == "HEX": _style_hex_cell(c); continue
                elif bg: c.fill = _fill(bg)
                c.alignment = _left(); c.border = _border(); c.font = _font(size=9)
            dr += 1
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Вспомогательные генераторы листов SRT
# ---------------------------------------------------------------------------

def _srt_sheet(wb, name: str, title_str: str, cols: list[str], ctypes: list[str],
               rows: list[_Row], val_fn, bg_fn=None, section_color=HEADER2) -> None:
    ws = wb.create_sheet(name)
    _title(ws, title_str, len(cols), section_color)
    _headers(ws, cols, row=2)
    _col_labels(ws, ctypes, row=3)
    for dr, row in enumerate(rows, 4):
        vals = val_fn(row)
        bg   = bg_fn(row) if bg_fn else (GRAY if dr % 2 == 0 else None)
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=dr, column=col, value=v)
            tp = ctypes[col-1]
            if tp == "ATTR": c.fill = _fill(YELLOW)
            elif tp == "HEX": _style_hex_cell(c); continue
            elif bg: c.fill = _fill(bg)
            c.alignment = _left(); c.border = _border(); c.font = _font(size=9)
    _auto_width(ws)


def _sheet_pos_data(wb, packets):
    rows = _collect(packets, 16)
    cols = ["PKT#","SDR#","SR#",
            "NTM","LAT","LONG",
            "VLD","FIX","MV","BB","CS_WGS84","ALTE",
            "SPD_kmh","DIR_deg","ODM_km","DIN","SRC","ALT_m","SRCD",
            "HEX_SRD"]
    ct   = ["","","",
            "ATTR","ATTR","ATTR",
            "ATTR","ATTR","ATTR","ATTR","ATTR","ATTR",
            "ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR",
            "HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [
            r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
            d.get("NTM"), d.get("LAT"), d.get("LONG"),
            d.get("VLD"), d.get("FIX"), d.get("MV"), d.get("BB"), d.get("CS_WGS84"), d.get("ALTE"),
            d.get("SPD_kmh"), d.get("DIR_deg"), d.get("ODM_km"),
            d.get("DIN"), d.get("SRC"), d.get("ALT_m"), d.get("SRCD"),
            r.rd.subrecord.to_bytes().hex().upper(),
        ]
    def _bg(r: _Row):
        d = r.rd.subrecord.to_dict()
        return GREEN if d.get("VLD") == 1 else GRAY
    _srt_sheet(wb, "SRT_POS_DATA", "EGTS_SR_POS_DATA (SRT=16) — Данные позиции", cols, ct, rows, _vals, _bg)


def _sheet_ext_pos(wb, packets):
    rows = _collect(packets, 17)
    cols = ["PKT#","SDR#","SR#","VDOP","HDOP","PDOP","SAT","NS","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("VDOP",""), d.get("HDOP",""), d.get("PDOP",""),
                d.get("SAT",""), d.get("NS",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_EXT_POS", "EGTS_SR_EXT_POS_DATA (SRT=17) — Доп. данные позиции", cols, ct, rows, _vals)


def _sheet_state(wb, packets):
    rows = _collect(packets, {20, 21})
    cols = ["PKT#","SDR#","SR#","SRT","ST","ST_desc","MPSV_V","BBV_V","IBV_V","NMS","IBU","BBU","HEX_SRD"]
    ct   = ["","","","","ATTR","","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1, r.rd.srt,
                d.get("ST"), d.get("ST_desc"),
                d.get("MPSV_V"), d.get("BBV_V"), d.get("IBV_V"),
                d.get("NMS"), d.get("IBU"), d.get("BBU"),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_STATE", "EGTS_SR_STATE_DATA (SRT=20/21) — Состояние терминала", cols, ct, rows, _vals)


def _sheet_liquid(wb, packets):
    rows = _collect(packets, 27)
    cols = ["PKT#","SDR#","SR#","LLSEF","LLSVU","LLSVU_desc","RDF","LLSN","MADDR","LLSD","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("LLSEF"), d.get("LLSVU"), d.get("LLSVU_desc"),
                d.get("RDF"), d.get("LLSN"), d.get("MADDR"), d.get("LLSD"),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_LIQUID", "EGTS_SR_LIQUID_LEVEL_SENSOR (SRT=27) — Уровень топлива", cols, ct, rows, _vals)


def _sheet_ad_sensors(wb, packets):
    rows = _collect(packets, 18)
    cols = (["PKT#","SDR#","SR#","DOUT"] +
            [f"ADIO{i}" for i in range(1,9)] +
            [f"ANS{i}"  for i in range(1,9)] +
            ["HEX_SRD"])
    ct   = (["","","","ATTR"] +
            ["ATTR"]*8 + ["ATTR"]*8 +
            ["HEX"])
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return ([r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1, d.get("DOUT","")]
                + [d.get(f"ADIO{i}","") for i in range(1,9)]
                + [d.get(f"ANS{i}","")  for i in range(1,9)]
                + [r.rd.subrecord.to_bytes().hex().upper()])
    _srt_sheet(wb, "SRT_AD_SENSORS", "EGTS_SR_AD_SENSORS_DATA (SRT=18) — Датчики", cols, ct, rows, _vals)


def _sheet_abs_cntr(wb, packets):
    rows = _collect(packets, 25)
    cols = ["PKT#","SDR#","SR#","ACN","ACV","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("ACN"), d.get("ACV"),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_ABS_CNTR", "EGTS_SR_ABS_CNTR_DATA (SRT=25) — Абсолютные счётчики", cols, ct, rows, _vals)


def _sheet_term_ident(wb, packets):
    rows = _collect(packets, 1)
    cols = ["PKT#","SDR#","SR#","TID","IMEI","IMSI","LNGC","NID","HDID","BS","MN","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("TID",""), d.get("IMEI",""), d.get("IMSI",""),
                d.get("LNGC",""), d.get("NID",""), d.get("HDID",""),
                d.get("BS",""), d.get("MN",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_TERM_IDENT", "EGTS_SR_TERM_IDENTITY (SRT=1) — Идентификация терминала", cols, ct, rows, _vals)


def _sheet_srt200(wb, packets):
    rows = _collect(packets, 200)
    cols = ["PKT#","SDR#","SR#","X_mm","Y_mm","Z_mm","quality","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("X_mm",""), d.get("Y_mm",""), d.get("Z_mm",""), d.get("quality",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_200", "EGTS_SR_CUSTOM_SRT200 (SRT=200) — RTLS Extended Position",
               cols, ct, rows, _vals, section_color="375623")


def _sheet_srt201(wb, packets):
    rows = _collect(packets, 201)
    cols = ["PKT#","SDR#","SR#","temperature_01C","temperature_C","vibration","pressure_hPa","sensor_flags","HEX_SRD"]
    ct   = ["","","","ATTR","HEX","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("temperature_01C",""), d.get("temperature_C",""),
                d.get("vibration",""), d.get("pressure_hPa",""), d.get("sensor_flags",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_201", "EGTS_SR_CUSTOM_SRT201 (SRT=201) — RTLS Sensor Data",
               cols, ct, rows, _vals, section_color="375623")


def _sheet_srt202(wb, packets):
    rows = _collect(packets, 202)
    cols = ["PKT#","SDR#","SR#","tag_id","zone_id","group_id","rssi_dBm","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("tag_id",""), d.get("zone_id",""), d.get("group_id",""), d.get("rssi_dBm",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_202", "EGTS_SR_CUSTOM_SRT202 (SRT=202) — RTLS Tag Identity",
               cols, ct, rows, _vals, section_color="375623")


def _sheet_srt203(wb, packets):
    rows = _collect(packets, 203)
    cols = ["PKT#","SDR#","SR#","event_type","event_desc","zone_id","event_time","event_flags","HEX_SRD"]
    ct   = ["","","","ATTR","","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("event_type",""), d.get("event_desc",""),
                d.get("zone_id",""), d.get("event_time",""), d.get("event_flags",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_203", "EGTS_SR_CUSTOM_SRT203 (SRT=203) — RTLS Event Data",
               cols, ct, rows, _vals, section_color="375623")


def _sheet_srt204(wb, packets):
    """IMU / Inertial data sheet (SRT=204) — orientation, accel/gyro, vibration, EKF + map-match outputs."""
    rows = _collect(packets, 204)
    cols = ["PKT#", "SDR#", "SR#",
            "heading_deg", "roll_deg", "pitch_deg", "heading_accuracy_deg",
            "accel_x", "accel_y", "accel_z",
            "gyro_x", "gyro_y", "gyro_z",
            "vibration_rms", "vibration_peak", "dominant_freq_hz", "filter_type",
            "ekf_confidence", "cov_trace",
            "road_segment_id", "matched_lat", "matched_lon", "snap_confidence",
            "flags", "timestamp", "HEX_SRD"]
    ct = ["", "", "",
          "ATTR", "ATTR", "ATTR", "ATTR",
          "ATTR", "ATTR", "ATTR",
          "ATTR", "ATTR", "ATTR",
          "ATTR", "ATTR", "ATTR", "ATTR",
          "ATTR", "ATTR",
          "ATTR", "ATTR", "ATTR", "ATTR",
          "ATTR", "ATTR", "HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("heading_deg", ""), d.get("roll_deg", ""), d.get("pitch_deg", ""),
                d.get("heading_accuracy_deg", ""),
                d.get("accel_x", ""), d.get("accel_y", ""), d.get("accel_z", ""),
                d.get("gyro_x", ""), d.get("gyro_y", ""), d.get("gyro_z", ""),
                d.get("vibration_rms", ""), d.get("vibration_peak", ""),
                d.get("dominant_freq_hz", ""), d.get("filter_type", ""),
                d.get("ekf_confidence", ""), d.get("cov_trace", ""),
                d.get("road_segment_id", ""), d.get("matched_lat", ""),
                d.get("matched_lon", ""), d.get("snap_confidence", ""),
                d.get("flags", ""), d.get("timestamp", ""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_204", "EGTS_SR_CUSTOM_SRT204 (SRT=204) — IMU + Inertial + Fusion outputs (discussion 09/13-16)",
               cols, ct, rows, _vals, section_color="843C0C")


def _sheet_srt205(wb, packets):
    """LBS data sheet (SRT=205, discussion 18) — base stations + TA/RSSI for road graph matching."""
    rows = _collect(packets, 205)
    cols = ["PKT#","SDR#","SR#","serving_cell_id","lac_tac","mcc","mnc","rssi_dbm","timing_advance",
            "bs_lat","bs_lon","raw_lbs_lat","raw_lbs_lon","neighbors","lbs_quality","technology","HEX_SRD"]
    ct   = ["","","","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR",
            "ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","ATTR","HEX"]
    def _vals(r: _Row):
        d = r.rd.subrecord.to_dict()
        neigh = d.get("neighbors", [])
        neigh_str = ";".join([f"{n.get('cell_id','')}:{n.get('rssi_dbm','')}" for n in neigh]) if neigh else ""
        return [r.pkt_idx+1, r.sdr_idx+1, r.sr_idx+1,
                d.get("serving_cell_id",""), d.get("lac_tac",""), d.get("mcc",""), d.get("mnc",""),
                d.get("rssi_dbm",""), d.get("timing_advance",""),
                d.get("bs_lat",""), d.get("bs_lon",""),
                d.get("raw_lbs_lat",""), d.get("raw_lbs_lon",""),
                neigh_str, d.get("lbs_quality",""), d.get("technology",""),
                r.rd.subrecord.to_bytes().hex().upper()]
    _srt_sheet(wb, "SRT_205", "EGTS_SR_CUSTOM_SRT205 (SRT=205) — LBS (base stations) for road graph (discussion 18)",
               cols, ct, rows, _vals, section_color="38761D")


def _sheet_raw(wb, packets):
    ws = wb.create_sheet("SRT_RAW")
    cols = ["PKT#","SDR#","SR#","SRT","SRT_name","SRL","raw_hex","note"]
    _title(ws, "Нераспознанные подзаписи (SRT_RAW)", len(cols))
    _headers(ws, cols, row=2)
    dr = 3
    for pi, pkt in enumerate(packets):
        body = pkt.body if isinstance(pkt.body, list) else []
        for si, sdr in enumerate(body):
            for ri, rd in enumerate(sdr.record_data):
                if not isinstance(rd.subrecord, SrRaw):
                    continue
                d = rd.subrecord.to_dict()
                vals = [pi+1, si+1, ri+1, rd.srt,
                        SRT_NAMES.get(rd.srt, f"SRT_{rd.srt}"), rd.srl,
                        d.get("raw_hex",""), d.get("note","")]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=dr, column=col, value=v)
                    c.fill = _fill(RED)
                    c.alignment = _left()
                    c.border = _border()
                    c.font = _font(size=9)
                dr += 1
    _auto_width(ws)


# ---------------------------------------------------------------------------
# ENCODE: workbook → binary packets
# ---------------------------------------------------------------------------

def encode_from_workbook(xlsx_path: str) -> list[EGTSPacket]:
    """
    Читает xlsx, восстанавливает пакеты из ATTR-колонок каждого листа.
    Возвращает список EGTSPacket готовых к to_bytes().
    """
    wb = openpyxl.load_workbook(xlsx_path)

    # Собираем атрибуты по ключу (pkt_idx, sdr_idx, sr_idx)
    pos_attrs    = _read_srt_sheet(wb, "SRT_POS_DATA",   ["PKT#","SDR#","SR#","NTM","LAT","LONG","VLD","FIX","MV","BB","CS_WGS84","ALTE","SPD_kmh","DIR_deg","ODM_km","DIN","SRC","ALT_m","SRCD"])
    state_attrs  = _read_srt_sheet(wb, "SRT_STATE",      ["PKT#","SDR#","SR#","SRT","ST","MPSV_V","BBV_V","IBV_V","NMS","IBU","BBU"])
    liquid_attrs = _read_srt_sheet(wb, "SRT_LIQUID",     ["PKT#","SDR#","SR#","LLSEF","LLSVU","RDF","LLSN","MADDR","LLSD"])
    cntr_attrs   = _read_srt_sheet(wb, "SRT_ABS_CNTR",  ["PKT#","SDR#","SR#","ACN","ACV"])
    srt200_attrs = _read_srt_sheet(wb, "SRT_200",        ["PKT#","SDR#","SR#","X_mm","Y_mm","Z_mm","quality"])
    srt201_attrs = _read_srt_sheet(wb, "SRT_201",        ["PKT#","SDR#","SR#","temperature_01C","vibration","pressure_hPa","sensor_flags"])
    srt202_attrs = _read_srt_sheet(wb, "SRT_202",        ["PKT#","SDR#","SR#","tag_id","zone_id","group_id","rssi_dBm"])
    srt203_attrs = _read_srt_sheet(wb, "SRT_203",        ["PKT#","SDR#","SR#","event_type","zone_id","event_time","event_flags"])
    srt205_attrs = _read_srt_sheet(wb, "SRT_205",        ["PKT#","SDR#","SR#","serving_cell_id","lac_tac","mcc","mnc","rssi_dbm","timing_advance",
                                                          "bs_lat","bs_lon","raw_lbs_lat","raw_lbs_lon","neighbors","lbs_quality","technology"])  # LBS (18)

    # Читаем заголовки и SDR-метаданные
    hdr_attrs = _read_srt_sheet(wb, "HEADER",
        ["PKT#","PRV","SKID","PRF","RTE","ENA","CMP","PR","HE","PID","PT","PRA","RCA","TTL"])
    sdr_attrs = _read_srt_sheet(wb, "SDR",
        ["PKT#","SDR#","RN","SSOD","RSOD","GRP","RPP","TMFE","EVFE","OBFE","OID","EVID","TM","SST","RST"])

    # Определяем количество пакетов
    pkt_count = max((int(r.get("PKT#", 0)) for r in hdr_attrs), default=0)
    packets: list[EGTSPacket] = []

    for pi in range(1, pkt_count + 1):
        from egts.models import Header
        from egts.codec  import EGTSPacket, ServiceDataRecord, RecordData

        # Заголовок
        hr = next((r for r in hdr_attrs if int(r.get("PKT#", 0)) == pi), {})
        hdr = Header()
        hdr.protocol_version  = int(hr.get("PRV", 1))
        hdr.security_key_id   = int(hr.get("SKID", 0))
        hdr.prefix            = str(hr.get("PRF", "00"))
        hdr.route             = str(hr.get("RTE", "0"))
        hdr.encryption_alg    = str(hr.get("ENA", "00"))
        hdr.compression       = str(hr.get("CMP", "0"))
        hdr.priority          = str(hr.get("PR", "00"))
        hdr.header_encoding   = int(hr.get("HE", 0))
        hdr.packet_id         = int(hr.get("PID", 0))
        hdr.packet_type       = int(hr.get("PT", 1))
        hdr.peer_address      = int(hr.get("PRA") or 0)
        hdr.recipient_address = int(hr.get("RCA") or 0)
        hdr.time_to_live      = int(hr.get("TTL") or 0)

        # SDR
        sdrs: list[ServiceDataRecord] = []
        sdr_rows = [r for r in sdr_attrs if int(r.get("PKT#", 0)) == pi]

        for si_row in sdr_rows:
            si = int(si_row.get("SDR#", 1))
            sdr = ServiceDataRecord(
                record_number=int(si_row.get("RN", 0)),
                ssod=str(si_row.get("SSOD", "0")),
                rsod=str(si_row.get("RSOD", "0")),
                grp=str(si_row.get("GRP", "0")),
                rpp=str(si_row.get("RPP", "00")),
                tmfe=str(si_row.get("TMFE", "0")),
                evfe=str(si_row.get("EVFE", "0")),
                obfe=str(si_row.get("OBFE", "0")),
                source_service_type=int(si_row.get("SST", 2)),
                recipient_service_type=int(si_row.get("RST", 2)),
            )

            # Собираем подзаписи из SRT-листов
            rd_list = []
            for srt_code, attr_list, Model, builder in [
                (16, pos_attrs,    SrPosData,           _build_pos_data),
                (17, [],           SrExtPosData,        None),
                (21, state_attrs,  SrStateData,         _build_state_data),
                (27, liquid_attrs, SrLiquidLevelSensor, _build_liquid),
                (25, cntr_attrs,   SrAbsCntrData,       _build_abs_cntr),
                (200, srt200_attrs, SrCustom200,        _build_srt200),
                (201, srt201_attrs, SrCustom201,        _build_srt201),
                (202, srt202_attrs, SrCustom202,        _build_srt202),
                (203, srt203_attrs, SrCustom203,        _build_srt203),
                (205, srt205_attrs, SrCustom205,        _build_srt205),  # LBS (18)
            ]:
                if builder is None:
                    continue
                for r in attr_list:
                    if int(r.get("PKT#", 0)) == pi and int(r.get("SDR#", 0)) == si:
                        subrecord = builder(r)
                        raw_b = subrecord.to_bytes()
                        rd_list.append(RecordData(srt=srt_code, srl=len(raw_b), subrecord=subrecord))

            rd_list.sort(key=lambda x: x.srt)
            sdr.record_data = rd_list
            sdrs.append(sdr)

        pkt = EGTSPacket(header=hdr, body=sdrs)
        packets.append(pkt)

    return packets


def _read_srt_sheet(wb: openpyxl.Workbook, sheet_name: str, col_names: list[str]) -> list[dict]:
    """Читает лист по именам колонок из строки 2. Возвращает список dict."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    # Ищем строку заголовка (строка 2)
    col_idx: dict[str, int] = {}
    for cell in ws[2]:
        if cell.value in col_names:
            col_idx[cell.value] = cell.column
    rows = []
    # Данные начинаются с 4-й строки (строка 3 — ATTR/HEX labels)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if all(v is None for v in row):
            continue
        d = {}
        for name, colidx in col_idx.items():
            d[name] = row[colidx - 1]
        rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# Построители подзаписей из dict атрибутов
# ---------------------------------------------------------------------------

def _safe_int(v, default=0) -> int:
    try:
        return int(v) if v is not None and v != "" else default
    except (TypeError, ValueError):
        return default


def _safe_float(v, default=0.0) -> float:
    try:
        return float(v) if v is not None and v != "" else default
    except (TypeError, ValueError):
        return default


def _build_pos_data(r: dict) -> SrPosData:
    from datetime import datetime, timezone
    obj = SrPosData()
    ntm_str = r.get("NTM")
    if ntm_str:
        try:
            obj.navigation_time = datetime.fromisoformat(str(ntm_str))
            if obj.navigation_time.tzinfo is None:
                obj.navigation_time = obj.navigation_time.replace(tzinfo=timezone.utc)
        except Exception:
            pass
    obj.latitude       = _safe_float(r.get("LAT"))
    obj.longitude      = _safe_float(r.get("LONG"))
    obj.valid          = bool(_safe_int(r.get("VLD"), 1))
    obj.fix_3d         = bool(_safe_int(r.get("FIX")))
    obj.in_motion      = bool(_safe_int(r.get("MV")))
    obj.black_box      = bool(_safe_int(r.get("BB")))
    obj.cs_wgs84       = bool(_safe_int(r.get("CS_WGS84"), 1))
    obj.altitude_exists = bool(_safe_int(r.get("ALTE")))
    obj.speed_kmh      = _safe_float(r.get("SPD_kmh"))
    obj.direction_deg  = _safe_int(r.get("DIR_deg"))
    obj.odometer_km    = _safe_float(r.get("ODM_km"))
    obj.digital_inputs = _safe_int(r.get("DIN"))
    obj.source         = _safe_int(r.get("SRC"))
    obj.altitude_m     = _safe_int(r.get("ALT_m"))
    obj.source_data    = _safe_int(r.get("SRCD"))
    return obj


def _build_state_data(r: dict) -> SrStateData:
    obj = SrStateData()
    obj.state = _safe_int(r.get("ST"))
    obj.mpsv  = int(round(_safe_float(r.get("MPSV_V")) * 10))
    obj.bbv   = int(round(_safe_float(r.get("BBV_V"))  * 10))
    obj.ibv   = int(round(_safe_float(r.get("IBV_V"))  * 10))
    obj.nms   = bool(_safe_int(r.get("NMS")))
    obj.ibu   = bool(_safe_int(r.get("IBU")))
    obj.bbu   = bool(_safe_int(r.get("BBU")))
    return obj


def _build_liquid(r: dict) -> SrLiquidLevelSensor:
    obj = SrLiquidLevelSensor()
    obj.error_flag  = bool(_safe_int(r.get("LLSEF")))
    obj.value_unit  = str(r.get("LLSVU", "00"))
    obj.raw_data    = bool(_safe_int(r.get("RDF")))
    obj.sensor_num  = _safe_int(r.get("LLSN"))
    obj.module_addr = _safe_int(r.get("MADDR"))
    obj.sensor_data = _safe_int(r.get("LLSD"))
    return obj


def _build_abs_cntr(r: dict) -> SrAbsCntrData:
    return SrAbsCntrData(
        channel_number=_safe_int(r.get("ACN")),
        counter_value=_safe_int(r.get("ACV")),
    )


def _build_srt200(r: dict) -> SrCustom200:
    return SrCustom200(x=_safe_int(r.get("X_mm")), y=_safe_int(r.get("Y_mm")),
                       z=_safe_int(r.get("Z_mm")), quality=_safe_int(r.get("quality")))


def _build_srt201(r: dict) -> SrCustom201:
    return SrCustom201(
        temperature_01c=_safe_int(r.get("temperature_01C")),
        vibration=_safe_int(r.get("vibration")),
        pressure_hpa=_safe_int(r.get("pressure_hPa")),
        sensor_flags=_safe_int(r.get("sensor_flags")),
    )


def _build_srt202(r: dict) -> SrCustom202:
    return SrCustom202(
        tag_id=_safe_int(r.get("tag_id")), zone_id=_safe_int(r.get("zone_id")),
        group_id=_safe_int(r.get("group_id")), rssi=_safe_int(r.get("rssi_dBm")),
    )


def _build_srt203(r: dict) -> SrCustom203:
    from datetime import datetime, timezone
    from egts.const import EPOCH
    obj = SrCustom203()
    obj.event_type  = _safe_int(r.get("event_type"))
    obj.zone_id     = _safe_int(r.get("zone_id"))
    obj.event_flags = _safe_int(r.get("event_flags"))
    ts_str = r.get("event_time")
    if ts_str:
        try:
            dt = datetime.fromisoformat(str(ts_str))
            obj.event_time = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except Exception:
            pass
    return obj


def _build_srt205(r: dict) -> SrCustom205:
    """Build LBS subrecord (discussion 18)."""
    obj = SrCustom205()
    obj.serving_cell_id = _safe_int(r.get("serving_cell_id"))
    obj.lac_tac = _safe_int(r.get("lac_tac"))
    obj.mcc = _safe_int(r.get("mcc"))
    obj.mnc = _safe_int(r.get("mnc"))
    obj.rssi_dbm = _safe_int(r.get("rssi_dbm"))
    obj.timing_advance = _safe_int(r.get("timing_advance"))
    obj.bs_lat = _safe_float(r.get("bs_lat"))
    obj.bs_lon = _safe_float(r.get("bs_lon"))
    obj.raw_lbs_lat = _safe_float(r.get("raw_lbs_lat"))
    obj.raw_lbs_lon = _safe_float(r.get("raw_lbs_lon"))
    obj.lbs_quality = _safe_int(r.get("lbs_quality"))
    obj.technology = _safe_int(r.get("technology"))
    # neighbors parsing omitted for simplicity (string "id:rssi;...")
    return obj


# ---------------------------------------------------------------------------
# Update HEX columns in workbook after encode
# ---------------------------------------------------------------------------

def _update_hex_in_workbook(wb: openpyxl.Workbook, packets: list[EGTSPacket]) -> None:
    """Обновляет все HEX-колонки в уже открытой книге."""
    # PACKETS — HEX_PACKET (последняя колонка)
    if "PACKETS" in wb.sheetnames:
        ws = wb["PACKETS"]
        hdr_row = [c.value for c in ws[2]]
        if "HEX_PACKET" in hdr_row:
            col = hdr_row.index("HEX_PACKET") + 1
            for i, pkt in enumerate(packets):
                raw = pkt.to_bytes()
                c = ws.cell(row=i + 4, column=col, value=raw.hex().upper())
                _style_hex_cell(c)

    # SRT-листы — HEX_SRD
    _srt_map = {
        "SRT_POS_DATA":  (16, "PKT#", "SDR#", "SR#"),
        "SRT_EXT_POS":   (17, "PKT#", "SDR#", "SR#"),
        "SRT_STATE":     ({20,21}, "PKT#", "SDR#", "SR#"),
        "SRT_LIQUID":    (27, "PKT#", "SDR#", "SR#"),
        "SRT_ABS_CNTR":  (25, "PKT#", "SDR#", "SR#"),
        "SRT_AD_SENSORS":(18, "PKT#", "SDR#", "SR#"),
        "SRT_200":       (200, "PKT#", "SDR#", "SR#"),
        "SRT_201":       (201, "PKT#", "SDR#", "SR#"),
        "SRT_202":       (202, "PKT#", "SDR#", "SR#"),
        "SRT_203":       (203, "PKT#", "SDR#", "SR#"),
    }
    for sheet_name, (srt_f, *_) in _srt_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        hdr_row = [c.value for c in ws[2]]
        if "HEX_SRD" not in hdr_row:
            continue
        hex_col = hdr_row.index("HEX_SRD") + 1
        pi_col  = hdr_row.index("PKT#")  + 1 if "PKT#"  in hdr_row else None
        si_col  = hdr_row.index("SDR#")  + 1 if "SDR#"  in hdr_row else None
        ri_col  = hdr_row.index("SR#")   + 1 if "SR#"   in hdr_row else None
        if not (pi_col and si_col and ri_col):
            continue
        for row in ws.iter_rows(min_row=4):
            pi = int(row[pi_col-1].value or 0)
            si = int(row[si_col-1].value or 0)
            ri = int(row[ri_col-1].value or 0)
            if pi == 0:
                continue
            try:
                pkt  = packets[pi - 1]
                body = pkt.body if isinstance(pkt.body, list) else []
                sdr  = body[si - 1]
                rd   = sdr.record_data[ri - 1]
                hex_val = rd.subrecord.to_bytes().hex().upper()
                c = row[hex_col - 1]
                c.value = hex_val
                _style_hex_cell(c)
            except (IndexError, AttributeError):
                pass


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _load_csv(path: str) -> list[EGTSPacket]:
    pkts: list[EGTSPacket] = []
    with open(path, encoding="utf-8") as f:
        for line_no, line in enumerate(f):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            hex_str = line.split(",")[0].strip()
            try:
                pkts.extend(parse_stream(bytes.fromhex(hex_str)))
            except Exception as e:
                print(f"  Line {line_no+1}: {e}")
    return pkts


def main() -> None:
    args = sys.argv[1:]
    if len(args) < 2:
        print(__doc__)
        sys.exit(0)

    cmd = args[0].lower()

    if cmd == "decode":
        if args[1] == "--hex":
            if len(args) < 3:
                print("Usage: decode --hex <HEX> [output.xlsx]")
                sys.exit(1)
            pkts = parse_stream(bytes.fromhex(args[2].replace(" ", "")))
            out  = args[3] if len(args) > 3 else "egts_decoded.xlsx"
        else:
            pkts = _load_csv(args[1])
            out  = args[2] if len(args) > 2 else "egts_decoded.xlsx"
        print(f"Decoded: {len(pkts)} packets")
        wb = decode_to_workbook(pkts)
        wb.save(out)
        print(f"Saved: {out}")

    elif cmd == "encode":
        xlsx_path = args[1]
        bin_path  = args[2] if len(args) > 2 else xlsx_path.replace(".xlsx", ".bin")
        print(f"Encoding from: {xlsx_path}")
        packets = encode_from_workbook(xlsx_path)
        print(f"Rebuilt: {len(packets)} packets")

        # Пересчёт и запись bin
        with open(bin_path, "wb") as f:
            for p in packets:
                raw = p.to_bytes()
                f.write(raw)
                print(f"  PKT #{p.header.packet_id}: {len(raw)} bytes  HEX: {raw.hex().upper()[:64]}...")
        print(f"Binary: {bin_path}")

        # Обновление HEX в xlsx
        wb = openpyxl.load_workbook(xlsx_path)
        _update_hex_in_workbook(wb, packets)
        wb.save(xlsx_path)
        print(f"HEX columns updated: {xlsx_path}")

    elif cmd == "roundtrip":
        xlsx_path = args[1]
        print(f"Roundtrip: {xlsx_path}")
        packets = encode_from_workbook(xlsx_path)
        print(f"Re-encoded: {len(packets)} packets")
        wb = openpyxl.load_workbook(xlsx_path)
        _update_hex_in_workbook(wb, packets)
        wb.save(xlsx_path)
        print(f"Updated: {xlsx_path}")
        for i, p in enumerate(packets):
            raw = p.to_bytes()
            print(f"  PKT #{i+1}: {raw.hex().upper()[:80]}")

    else:
        print(f"Unknown command: {cmd}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
